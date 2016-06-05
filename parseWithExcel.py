import bs4,os
import openpyxl
from openpyxl.cell import get_column_letter , column_index_from_string


def parse():
    
    listall = []
    #set working directory as folder which contains the files
    os.chdir('D:\\sem4res\\')
    for filename in os.listdir(os.getcwd()):
        
    #create a soup object of html file
    #path = 'D:\\newexam\\show_result04c5.html'
        
        soup = bs4.BeautifulSoup(open(filename),'html.parser')
        #print(filename)
        #start the parsing
        alltable = soup.select('table')

        name = alltable[1]
        score = alltable[4]

        ANS = ''
        #find the name and roll no
        nametext = name.text
        split_name_text = nametext.split('\n')
        #now 4 contains name and 6 contains roll no
        listname = split_name_text[4].split(' ')
        roll = split_name_text[6]
        #now 0 and 1 contains value in caps
        n1 = listname[0]
        n2 = listname[1]
        first_name=''
        last_name=''
        for char in n1:
            if char.isupper()==True:
                first_name = first_name + char

        for char in n2:
            if char.isupper()==True:
                last_name = last_name + char
        if last_name == '':
            last_name = '*'
        #now we need the roll number
        roll_no = roll[32]+roll[33]

        #NOW WE PARSE THE SCORE

        scoretext = score.text.replace('\xa0','')
        scoretext = scoretext.replace(' ','')
        listscore = scoretext.split('\n')
        #remove null values
        while True:
            try:
                listscore.remove('')
            except:
                break

        sc = ''
        for index in range(8,44,4):
            sc = sc + ' ' + listscore[index]

        ANS = roll_no + ' '+ first_name + ' ' + last_name + sc
        listall.append(ANS)
    listall.sort()
    return listall

def gradeToMarks(grade):
    if grade=='S':
        return 10
    elif grade=='A':
        return 9
    elif grade=='B':
        return 8
    elif grade=='C':
        return 7
    elif grade=='D':
        return 6
    elif grade=='E':
        return 5
    else:
        return 0

def saveInExcel(listAll):
    
    N = len(listAll)
    #NOW THE WORK OF SHEET YO
    wb = openpyxl.load_workbook ('D:\\sem42.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    row = 2
    while row<=(N+1):
        #get all the stuff for each studdent
        stuff = listAll[row-2].split(' ')
        print(stuff)
        s=0
        #now we traverse with column filling values
        for index in range(1,13):
            address = get_column_letter(index)+str(row)
            sheet[address]=stuff[index-1]
            if (index-1)>=3:
                s = s + gradeToMarks(stuff[index-1])
        address = 'M'+str(row)
        sheet[address]=s/9.0
        row = row + 1
    wb.save('D:\\SEM4.xlsx')


listAll = parse()
saveInExcel(listAll)

    
